import os
import telebot
from telebot import types
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import json
import time
import re
from telebot.apihelper import ApiTelegramException  # <-- Ð”ÐžÐ‘ÐÐ’Ð›Ð•ÐÐž

# ====== Ð¡Ð•ÐšÐ Ð•Ð¢ÐÐ«Ð• Ð”ÐÐÐÐ«Ð• Ð˜Ð— ÐŸÐ•Ð Ð•ÐœÐ•ÐÐÐ«Ð¥ ÐžÐšÐ Ð£Ð–Ð•ÐÐ˜Ð¯ ======
BOT_TOKEN = os.getenv('BOT_TOKEN')
ADMIN_ID = int(os.getenv('ADMIN_ID', '0'))

if not BOT_TOKEN:
    raise ValueError("BOT_TOKEN Ð½Ðµ Ð·Ð°Ð´Ð°Ð½ Ð² Ð¿ÐµÑ€ÐµÐ¼ÐµÐ½Ð½Ñ‹Ñ… Ð¾ÐºÑ€ÑƒÐ¶ÐµÐ½Ð¸Ñ!")
if not ADMIN_ID:
    raise ValueError("ADMIN_ID Ð½Ðµ Ð·Ð°Ð´Ð°Ð½ Ð² Ð¿ÐµÑ€ÐµÐ¼ÐµÐ½Ð½Ñ‹Ñ… Ð¾ÐºÑ€ÑƒÐ¶ÐµÐ½Ð¸Ñ!")
# =====================================================

# ====== ÐŸÐ£Ð‘Ð›Ð˜Ð§ÐÐ«Ð• Ð”ÐÐÐÐ«Ð• Ð¡ÐÐ›ÐžÐÐ ======
SALON_NAME = "Ð¡Ñ‚ÑƒÐ´Ð¸Ñ ÐºÑ€Ð°ÑÐ¾Ñ‚Ñ‹ â€œKÄ°VÄ°â€"
SALON_PHONE = "+7 (985) 699-17-77"
SALON_ADDRESS = "Ð¼. ÐŸÑÑ‚Ð½Ð¸Ñ†ÐºÐ¾Ðµ ÑˆÐ¾ÑÑÐµ, ÐÐ½Ð³ÐµÐ»Ð¾Ð² Ð¿ÐµÑ€ÐµÑƒÐ»Ð¾Ðº, Ð´Ð¾Ð¼ 2"
SALON_HOURS = "Ð‘ÐµÐ· Ð²Ñ‹Ñ…Ð¾Ð´Ð½Ñ‹Ñ… Ñ 10:00 Ð´Ð¾ 22:00"
SALON_BOOKING_URL = "https://n1610700.yclients.com"
SALON_TELEGRAM = "@kivi_mitino"
WEB_APP_URL = "https://48fill777.github.io/wheel-of-fortune/"
# ======================================

bot = telebot.TeleBot(BOT_TOKEN)

# Ð¤ÑƒÐ½ÐºÑ†Ð¸Ñ Ð±ÐµÐ·Ð¾Ð¿Ð°ÑÐ½Ð¾Ð¹ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐ¸ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹ (Ð¾Ð±Ñ€Ð°Ð±Ð°Ñ‚Ñ‹Ð²Ð°ÐµÑ‚ Ð±Ð»Ð¾ÐºÐ¸Ñ€Ð¾Ð²ÐºÑƒ Ð±Ð¾Ñ‚Ð°)
def safe_send_message(chat_id, text, **kwargs):
    try:
        bot.send_message(chat_id, text, **kwargs)
    except ApiTelegramException as e:
        if e.error_code == 403:
            # ÐŸÐ¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÑŒ Ð·Ð°Ð±Ð»Ð¾ÐºÐ¸Ñ€Ð¾Ð²Ð°Ð» Ð±Ð¾Ñ‚Ð° â€” Ð¿Ñ€Ð¾ÑÑ‚Ð¾ Ð¸Ð³Ð½Ð¾Ñ€Ð¸Ñ€ÑƒÐµÐ¼
            print(f"âš ï¸ ÐŸÐ¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÑŒ {chat_id} Ð·Ð°Ð±Ð»Ð¾ÐºÐ¸Ñ€Ð¾Ð²Ð°Ð» Ð±Ð¾Ñ‚Ð°, ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ Ð½Ðµ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¾")
        else:
            # Ð”Ñ€ÑƒÐ³Ð¸Ðµ Ð¾ÑˆÐ¸Ð±ÐºÐ¸ API (Ð½Ð°Ð¿Ñ€Ð¸Ð¼ÐµÑ€, ÑÐ»Ð¸ÑˆÐºÐ¾Ð¼ Ð¼Ð½Ð¾Ð³Ð¾ Ð·Ð°Ð¿Ñ€Ð¾ÑÐ¾Ð²) â€” Ð»Ð¾Ð³Ð¸Ñ€ÑƒÐµÐ¼ Ð¸ Ð½Ðµ Ð¿Ñ€ÐµÑ€Ñ‹Ð²Ð°ÐµÐ¼ Ñ€Ð°Ð±Ð¾Ñ‚Ñƒ
            print(f"âš ï¸ ÐžÑˆÐ¸Ð±ÐºÐ° Telegram API Ð¿Ñ€Ð¸ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐµ {chat_id}: {e}")
    except Exception as e:
        # ÐÐµÐ¿Ñ€ÐµÐ´Ð²Ð¸Ð´ÐµÐ½Ð½Ð°Ñ Ð¾ÑˆÐ¸Ð±ÐºÐ° â€” Ð¿Ñ€Ð¾Ð±Ñ€Ð°ÑÑ‹Ð²Ð°ÐµÐ¼ Ð´Ð°Ð»ÑŒÑˆÐµ, Ñ‡Ñ‚Ð¾Ð±Ñ‹ Ð²Ð½ÐµÑˆÐ½Ð¸Ð¹ Ñ†Ð¸ÐºÐ» Ð¿ÐµÑ€ÐµÐ·Ð°Ð¿ÑƒÑÑ‚Ð¸Ð» Ð±Ð¾Ñ‚Ð°
        print(f"âŒ ÐšÑ€Ð¸Ñ‚Ð¸Ñ‡ÐµÑÐºÐ°Ñ Ð¾ÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¸ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐµ {chat_id}: {e}")
        raise

# Ð¡Ð±Ñ€Ð°ÑÑ‹Ð²Ð°ÐµÐ¼ Ð²ÐµÐ±Ñ…ÑƒÐº (Ð²Ð°Ð¶Ð½Ð¾ Ð´Ð»Ñ polling)
bot.remove_webhook()
time.sleep(1)

EXCEL_FILE = 'clients_data.xlsx'

# Ð˜Ð½Ð¸Ñ†Ð¸Ð°Ð»Ð¸Ð·Ð°Ñ†Ð¸Ñ Excel
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws_clients = wb.active
        ws_clients.title = "ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"
        headers = ["telegram_id", "username", "full_name", "phone", "prize", "win_date", "is_used"]
        ws_clients.append(headers)
        wb.save(EXCEL_FILE)

init_excel()

def has_user_spun(telegram_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and int(row[0]) == telegram_id:
            return True
    return False

def add_spin_record(telegram_id, username, full_name, prize):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and int(row[0]) == telegram_id:
            return False
    ws.append([telegram_id, username, full_name, "", prize, datetime.now().isoformat(), 0])
    wb.save(EXCEL_FILE)
    return True

def update_phone(telegram_id, phone):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"]
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cell_value = row[0].value
        if cell_value is not None and int(cell_value) == telegram_id:
            ws.cell(row=i, column=4).value = phone
            wb.save(EXCEL_FILE)
            return True
    return False

def get_user_record(telegram_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and int(row[0]) == telegram_id:
            return i, row
    return None, None

def validate_phone(phone):
    phone = re.sub(r'\D', '', phone)
    return len(phone) in (10, 11)

def format_phone(phone):
    phone = re.sub(r'\D', '', phone)
    if len(phone) == 11:
        phone = phone[1:]
    return f"+7 ({phone[:3]}) {phone[3:6]}-{phone[6:8]}-{phone[8:]}"

@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.from_user.id
    spun = has_user_spun(user_id)
    url = WEB_APP_URL + ("?already_spun=1" if spun else "")
    print(f"[DEBUG] /start Ð´Ð»Ñ {user_id}, spun={spun}")

    # Reply-ÐºÐ½Ð¾Ð¿ÐºÐ° Ð´Ð»Ñ Ð¾Ñ‚ÐºÑ€Ñ‹Ñ‚Ð¸Ñ ÐºÐ¾Ð»ÐµÑÐ°
    markup_reply = types.ReplyKeyboardMarkup(resize_keyboard=True)
    web_app_button = types.KeyboardButton(
        text="ðŸŽ¡ ÐšÑ€ÑƒÑ‚Ð¸Ñ‚ÑŒ ÐºÐ¾Ð»ÐµÑÐ¾!",
        web_app=types.WebAppInfo(url=url)
    )
    markup_reply.add(web_app_button)

    # ÐŸÑ€Ð¸Ð²ÐµÑ‚ÑÑ‚Ð²ÐµÐ½Ð½Ð¾Ðµ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ
    safe_send_message(
        message.chat.id,
        f"ðŸŒŸ Ð”Ð¾Ð±Ñ€Ð¾ Ð¿Ð¾Ð¶Ð°Ð»Ð¾Ð²Ð°Ñ‚ÑŒ Ð² Ð¡Ñ‚ÑƒÐ´Ð¸ÑŽ ÐºÑ€Ð°ÑÐ¾Ñ‚Ñ‹ â€œKÄ°VÄ°â€! ðŸŒŸ\n\n"
        f"ÐœÑ‹ Ð´Ð°Ñ€Ð¸Ð¼ Ð¿Ð¾Ð´Ð°Ñ€ÐºÐ¸ ÐºÐ°Ð¶Ð´Ð¾Ð¼Ñƒ Ð½Ð¾Ð²Ð¾Ð¼Ñƒ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ñƒ!\n"
        f"ÐšÑ€ÑƒÑ‚Ð¸Ñ‚Ðµ ÐºÐ¾Ð»ÐµÑÐ¾ Ñ„Ð¾Ñ€Ñ‚ÑƒÐ½Ñ‹ Ð¸ Ð²Ñ‹Ð¸Ð³Ñ€Ñ‹Ð²Ð°Ð¹Ñ‚Ðµ:\n\n"
        f"ðŸ’… Ð”Ð¸Ð·Ð°Ð¹Ð½ Ð½Ð¾Ð³Ñ‚ÐµÐ¹\n"
        f"ðŸ§´ Ð¡ÐŸÐ Ð´Ð»Ñ Ñ€ÑƒÐº/Ð½Ð¾Ð³\n"
        f"ðŸ’° Ð¡ÐºÐ¸Ð´ÐºÐ° 10%\n"
        f"ðŸ’† ÐœÐ°ÑÑÐ°Ð¶ Ð²Ð¾Ñ€Ð¾Ñ‚Ð½Ð¸ÐºÐ¾Ð²Ð¾Ð¹ Ð·Ð¾Ð½Ñ‹\n"
        f"ðŸ’Ž Ð”ÐµÐ¿Ð¾Ð·Ð¸Ñ‚ 1 000 Ñ€ÑƒÐ±.\n"
        f"ðŸ‘‘ Ð”ÐµÐ¿Ð¾Ð·Ð¸Ñ‚ 10 000 Ñ€ÑƒÐ±.\n\n"
        f"ðŸŽ¯ Ð”Ð»Ñ Ð°ÐºÑ‚Ð¸Ð²Ð°Ñ†Ð¸Ð¸ Ð¿Ð¾Ð´Ð°Ñ€ÐºÐ° Ð¿Ð¾Ñ‚Ñ€ÐµÐ±ÑƒÐµÑ‚ÑÑ Ð½Ð¾Ð¼ÐµÑ€ Ñ‚ÐµÐ»ÐµÑ„Ð¾Ð½Ð°.\n"
        f"ÐžÐ±Ñ€Ð°Ñ‚Ð¸Ñ‚Ðµ Ð²Ð½Ð¸Ð¼Ð°Ð½Ð¸Ðµ: ÑƒÑ‡Ð°ÑÑ‚Ð²Ð¾Ð²Ð°Ñ‚ÑŒ Ð¼Ð¾Ð¶Ð½Ð¾ Ñ‚Ð¾Ð»ÑŒÐºÐ¾ Ð¾Ð´Ð¸Ð½ Ñ€Ð°Ð·!\n"
        f"ÐŸÐ¾Ð´Ð°Ñ€Ð¾Ðº Ð´ÐµÐ¹ÑÑ‚Ð²Ð¸Ñ‚ÐµÐ»ÐµÐ½ Ð² Ñ‚ÐµÑ‡ÐµÐ½Ð¸Ðµ 30 Ð´Ð½ÐµÐ¹.",
        reply_markup=markup_reply
    )

    # Inline-ÐºÐ½Ð¾Ð¿ÐºÐ¸ (ÐºÐ¾Ð½Ñ‚Ð°ÐºÑ‚Ñ‹, Ð·Ð°Ð¿Ð¸ÑÑŒ, Ð¼Ð¾Ð¹ Ð²Ñ‹Ð¸Ð³Ñ€Ñ‹Ñˆ)
    markup_inline = types.InlineKeyboardMarkup(row_width=2)
    btn_contacts = types.InlineKeyboardButton('ðŸ“ž ÐšÐ¾Ð½Ñ‚Ð°ÐºÑ‚Ñ‹', callback_data='contacts')
    btn_booking = types.InlineKeyboardButton('ðŸ“… Ð—Ð°Ð¿Ð¸ÑÐ°Ñ‚ÑŒÑÑ Ð¾Ð½Ð»Ð°Ð¹Ð½', url=SALON_BOOKING_URL)
    btn_prize = types.InlineKeyboardButton('ðŸŽ ÐœÐ¾Ð¹ Ð²Ñ‹Ð¸Ð³Ñ€Ñ‹Ñˆ', callback_data='my_prize')
    markup_inline.add(btn_contacts, btn_booking, btn_prize)

    safe_send_message(
        message.chat.id,
        "ÐÐ°ÑˆÐ¸ ÐºÐ¾Ð½Ñ‚Ð°ÐºÑ‚Ñ‹ Ð¸ Ð·Ð°Ð¿Ð¸ÑÑŒ:",
        reply_markup=markup_inline
    )

@bot.message_handler(content_types=['web_app_data'])
def handle_web_app_data(message):
    print(f"âœ… ÐŸÐžÐ›Ð£Ð§Ð•ÐÐ« WEB_APP_DATA: {message.web_app_data.data}")
    try:
        data = json.loads(message.web_app_data.data)
        prize_name = data['prize']
        user_id = message.from_user.id
        username = message.from_user.username or ""
        full_name = message.from_user.full_name

        if has_user_spun(user_id):
            safe_send_message(message.chat.id, "âŒ Ð’Ñ‹ ÑƒÐ¶Ðµ ÑƒÑ‡Ð°ÑÑ‚Ð²Ð¾Ð²Ð°Ð»Ð¸.")
            return

        if add_spin_record(user_id, username, full_name, prize_name):
            safe_send_message(ADMIN_ID, f"ðŸŽ‰ ÐÐ¾Ð²Ñ‹Ð¹ Ð²Ñ‹Ð¸Ð³Ñ€Ñ‹Ñˆ: {prize_name} Ð¾Ñ‚ {full_name} (@{username})")
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton('ðŸ“± ÐžÑÑ‚Ð°Ð²Ð¸Ñ‚ÑŒ Ð½Ð¾Ð¼ÐµÑ€', callback_data='enter_phone'))
            safe_send_message(
                message.chat.id,
                f"ðŸŽ‰ Ð’Ñ‹ Ð²Ñ‹Ð¸Ð³Ñ€Ð°Ð»Ð¸: {prize_name}!\n\nÐÐ°Ð¶Ð¼Ð¸Ñ‚Ðµ ÐºÐ½Ð¾Ð¿ÐºÑƒ, Ñ‡Ñ‚Ð¾Ð±Ñ‹ Ð¾ÑÑ‚Ð°Ð²Ð¸Ñ‚ÑŒ Ð½Ð¾Ð¼ÐµÑ€.",
                reply_markup=markup
            )
        else:
            safe_send_message(message.chat.id, "âŒ ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½Ð¸Ñ.")
    except Exception as e:
        print(f"[ERROR] {e}")

@bot.callback_query_handler(func=lambda call: call.data == 'enter_phone')
def phone_request(call):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add(types.KeyboardButton('ðŸ“± ÐžÑ‚Ð¿Ñ€Ð°Ð²Ð¸Ñ‚ÑŒ Ð½Ð¾Ð¼ÐµÑ€', request_contact=True))
    safe_send_message(call.message.chat.id, "ðŸ“± ÐžÑ‚Ð¿Ñ€Ð°Ð²ÑŒÑ‚Ðµ Ð½Ð¾Ð¼ÐµÑ€ Ñ‚ÐµÐ»ÐµÑ„Ð¾Ð½Ð°:", reply_markup=markup)
    bot.answer_callback_query(call.id)

@bot.message_handler(content_types=['contact'])
def handle_contact(message):
    phone = message.contact.phone_number
    formatted = format_phone(phone)
    if update_phone(message.from_user.id, formatted):
        _, record = get_user_record(message.from_user.id)
        prize = record[4] if record else "Ð¿Ñ€Ð¸Ð·"
        safe_send_message(ADMIN_ID, f"ðŸ“ž ÐŸÐ¾Ð»ÑƒÑ‡ÐµÐ½ Ð½Ð¾Ð¼ÐµÑ€: {formatted} (Ð¿Ñ€Ð¸Ð·: {prize})")
        safe_send_message(
            message.chat.id,
            f"âœ… Ð¡Ð¿Ð°ÑÐ¸Ð±Ð¾! Ð’Ð°Ñˆ Ð½Ð¾Ð¼ÐµÑ€ {formatted} ÑÐ¾Ñ…Ñ€Ð°Ð½Ñ‘Ð½. ÐÐ´Ð¼Ð¸Ð½Ð¸ÑÑ‚Ñ€Ð°Ñ‚Ð¾Ñ€ ÑÐ²ÑÐ¶ÐµÑ‚ÑÑ Ñ Ð²Ð°Ð¼Ð¸.",
            reply_markup=types.ReplyKeyboardRemove()
        )
    else:
        safe_send_message(message.chat.id, "âŒ ÐžÑˆÐ¸Ð±ÐºÐ°. ÐÐ°Ñ‡Ð½Ð¸Ñ‚Ðµ Ð·Ð°Ð½Ð¾Ð²Ð¾ /start")

@bot.message_handler(func=lambda m: m.text and m.text[0].isdigit())
def manual_phone(message):
    phone = message.text.strip()
    if validate_phone(phone):
        formatted = format_phone(phone)
        if update_phone(message.from_user.id, formatted):
            _, record = get_user_record(message.from_user.id)
            prize = record[4] if record else "Ð¿Ñ€Ð¸Ð·"
            safe_send_message(ADMIN_ID, f"ðŸ“ž ÐŸÐ¾Ð»ÑƒÑ‡ÐµÐ½ Ð½Ð¾Ð¼ÐµÑ€ (Ð²Ñ€ÑƒÑ‡Ð½ÑƒÑŽ): {formatted} (Ð¿Ñ€Ð¸Ð·: {prize})")
            safe_send_message(message.chat.id, f"âœ… Ð¡Ð¿Ð°ÑÐ¸Ð±Ð¾! ÐÐ¾Ð¼ÐµÑ€ {formatted} ÑÐ¾Ñ…Ñ€Ð°Ð½Ñ‘Ð½.")
        else:
            safe_send_message(message.chat.id, "âŒ Ð¡Ð½Ð°Ñ‡Ð°Ð»Ð° Ð½ÑƒÐ¶Ð½Ð¾ Ð²Ñ‹Ð¸Ð³Ñ€Ð°Ñ‚ÑŒ Ð¿Ñ€Ð¸Ð·. /start")
    else:
        safe_send_message(message.chat.id, "âŒ ÐÐµÐ²ÐµÑ€Ð½Ñ‹Ð¹ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚. ÐŸÑ€Ð¸Ð¼ÐµÑ€: +79991234567")

@bot.message_handler(commands=['my_prize'])
def my_prize_command(message):
    user_id = message.from_user.id
    _, record = get_user_record(user_id)
    print(f"[DEBUG] my_prize Ð´Ð»Ñ {user_id}, record={record}")
    if record:
        status = "âœ… ÐÐºÑ‚Ð¸Ð²Ð¸Ñ€Ð¾Ð²Ð°Ð½" if record[6] == 1 else "â³ ÐžÐ¶Ð¸Ð´Ð°ÐµÑ‚"
        safe_send_message(
            message.chat.id,
            f"ðŸŽ Ð’Ð°Ñˆ Ð¿Ñ€Ð¸Ð·: {record[4]}\nÐ¡Ñ‚Ð°Ñ‚ÑƒÑ: {status}"
        )
    else:
        safe_send_message(message.chat.id, "âŒ Ð’Ñ‹ ÐµÑ‰Ñ‘ Ð½Ðµ ÑƒÑ‡Ð°ÑÑ‚Ð²Ð¾Ð²Ð°Ð»Ð¸.")

@bot.callback_query_handler(func=lambda call: call.data == 'my_prize')
def my_prize_callback(call):
    my_prize_command(call.message)
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data == 'contacts')
def show_contacts(call):
    text = f"""
ðŸ“ž ÐšÐ¾Ð½Ñ‚Ð°ÐºÑ‚Ñ‹ ÑÑ‚ÑƒÐ´Ð¸Ð¸ â€œKÄ°VÄ°â€

ðŸ“ ÐÐ´Ñ€ÐµÑ: {SALON_ADDRESS}
ðŸ• Ð ÐµÐ¶Ð¸Ð¼ Ñ€Ð°Ð±Ð¾Ñ‚Ñ‹: {SALON_HOURS}
ðŸ“± Ð¢ÐµÐ»ÐµÑ„Ð¾Ð½: {SALON_PHONE}
ðŸ’¬ Telegram: {SALON_TELEGRAM}
    """
    safe_send_message(call.message.chat.id, text)
    bot.answer_callback_query(call.id)

# ÐÐ´Ð¼Ð¸Ð½-Ð¿Ð°Ð½ÐµÐ»ÑŒ (ÐºÐ¾Ð¼Ð°Ð½Ð´Ð° /admin)
@bot.message_handler(commands=['admin'])
def admin_panel(message):
    if message.from_user.id != ADMIN_ID:
        return
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton('ðŸ“Š Ð¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ°', callback_data='admin_stats'),
        types.InlineKeyboardButton('â³ ÐžÐ¶Ð¸Ð´Ð°ÑŽÑ‚ Ð½Ð¾Ð¼ÐµÑ€Ð°', callback_data='admin_no_phone'),
        types.InlineKeyboardButton('ðŸ“ž ÐžÐ¶Ð¸Ð´Ð°ÑŽÑ‚ ÑÐ²ÑÐ·Ð¸', callback_data='admin_pending'),
        types.InlineKeyboardButton('ðŸ“‹ Ð’ÑÐµ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ñ‹', callback_data='admin_all')
    )
    safe_send_message(message.chat.id, "ðŸ”§ ÐÐ”ÐœÐ˜Ð-ÐŸÐÐÐ•Ð›Ð¬", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == 'admin_stats')
def admin_stats(call):
    if call.from_user.id != ADMIN_ID:
        return
    wb = load_workbook(EXCEL_FILE)
    ws = wb["ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"]
    total = ws.max_row - 1
    with_phone = 0
    used = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3]:
            with_phone += 1
        if row[6] == 1:
            used += 1
    text = f"""
ðŸ“Š Ð¡Ð¢ÐÐ¢Ð˜Ð¡Ð¢Ð˜ÐšÐ

ðŸ‘¥ Ð’ÑÐµÐ³Ð¾ ÑƒÑ‡Ð°ÑÑ‚Ð½Ð¸ÐºÐ¾Ð²: {total}
ðŸ“ž ÐžÑÑ‚Ð°Ð²Ð¸Ð»Ð¸ Ð½Ð¾Ð¼ÐµÑ€: {with_phone}
âœ… ÐžÐ±ÑÐ»ÑƒÐ¶ÐµÐ½Ð¾: {used}
    """
    safe_send_message(call.message.chat.id, text)
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data == 'admin_no_phone')
def admin_no_phone(call):
    if call.from_user.id != ADMIN_ID:
        return
    wb = load_workbook(EXCEL_FILE)
    ws = wb["ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"]
    text = "â³ ÐžÐ–Ð˜Ð”ÐÐ®Ð¢ ÐÐžÐœÐ•Ð  Ð¢Ð•Ð›Ð•Ð¤ÐžÐÐ:\n\n"
    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[3]:
            found = True
            text += f"ðŸ‘¤ {row[2]} (@{row[1]})\nðŸ†” {row[0]}\nðŸŽ {row[4]}\nðŸ“… {row[5][:16]}\n\n"
    if not found:
        text = "âœ… Ð’ÑÐµ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ñ‹ Ð¾ÑÑ‚Ð°Ð²Ð¸Ð»Ð¸ Ð½Ð¾Ð¼ÐµÑ€."
    safe_send_message(call.message.chat.id, text)
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data == 'admin_pending')
def admin_pending(call):
    if call.from_user.id != ADMIN_ID:
        return
    wb = load_workbook(EXCEL_FILE)
    ws = wb["ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"]
    text = "â³ ÐžÐ–Ð˜Ð”ÐÐ®Ð¢ Ð¡Ð’Ð¯Ð—Ð˜ (ÐµÑÑ‚ÑŒ Ð½Ð¾Ð¼ÐµÑ€, Ð½Ðµ Ð¾Ð±ÑÐ»ÑƒÐ¶ÐµÐ½Ñ‹):\n\n"
    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] and row[6] == 0:
            found = True
            text += f"ðŸ‘¤ {row[2]} (@{row[1]})\nðŸ“ž {row[3]}\nðŸŽ {row[4]}\nðŸ“… {row[5][:16]}\n\n"
    if not found:
        text = "âœ… ÐÐµÑ‚ Ð¾Ð¶Ð¸Ð´Ð°ÑŽÑ‰Ð¸Ñ… ÑÐ²ÑÐ·Ð¸."
    safe_send_message(call.message.chat.id, text)
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data == 'admin_all')
def admin_all(call):
    if call.from_user.id != ADMIN_ID:
        return
    wb = load_workbook(EXCEL_FILE)
    ws = wb["ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"]
    text = "ðŸ“‹ Ð’Ð¡Ð• ÐšÐ›Ð˜Ð•ÐÐ¢Ð«:\n\n"
    for row in ws.iter_rows(min_row=2, values_only=True):
        phone = row[3] if row[3] else "Ð½Ðµ ÑƒÐºÐ°Ð·Ð°Ð½"
        status = "âœ…" if row[6] == 1 else "â³"
        text += f"{status} {row[2]} (@{row[1]}) ðŸ“ž {phone}\nðŸŽ {row[4]}\n\n"
    if ws.max_row == 1:
        text = "ÐŸÐ¾ÐºÐ° Ð½ÐµÑ‚ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð²."
    safe_send_message(call.message.chat.id, text)
    bot.answer_callback_query(call.id)

# ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ñ‡Ð¸Ðº Ð´Ð»Ñ Ð¾Ð±Ñ€Ð°Ñ‰ÐµÐ½Ð¸Ð¹ Ðº Ð°Ð´Ð¼Ð¸Ð½Ñƒ
@bot.message_handler(commands=['call_admin'])
def call_admin(message):
    safe_send_message(ADMIN_ID, f"ðŸ”” ÐšÐ»Ð¸ÐµÐ½Ñ‚ {message.from_user.full_name} (@{message.from_user.username}) Ð¿Ñ€Ð¾ÑÐ¸Ñ‚ Ð¿Ð¾Ð¼Ð¾Ñ‰Ð¸!")
    safe_send_message(message.chat.id, "âœ… Ð—Ð°Ð¿Ñ€Ð¾Ñ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½ Ð°Ð´Ð¼Ð¸Ð½Ð¸ÑÑ‚Ñ€Ð°Ñ‚Ð¾Ñ€Ñƒ.")

# Ð—Ð°Ð¿ÑƒÑÐº Ð±Ð¾Ñ‚Ð° Ñ Ð°Ð²Ñ‚Ð¾-Ð¿ÐµÑ€ÐµÐ·Ð°Ð¿ÑƒÑÐºÐ¾Ð¼
if __name__ == '__main__':
    print(f"ðŸš€ Ð‘Ð¾Ñ‚ Ð´Ð»Ñ ÑÐ°Ð»Ð¾Ð½Ð° '{SALON_NAME}' Ð·Ð°Ð¿ÑƒÑ‰ÐµÐ½!")
    print(f"ðŸ‘¤ ÐÐ´Ð¼Ð¸Ð½Ð¸ÑÑ‚Ñ€Ð°Ñ‚Ð¾Ñ€ ID: {ADMIN_ID}")
    print(f"ðŸ“ž Ð¢ÐµÐ»ÐµÑ„Ð¾Ð½: {SALON_PHONE}")
    print(f"ðŸ“ ÐÐ´Ñ€ÐµÑ: {SALON_ADDRESS}")
    print("ÐžÐ¶Ð¸Ð´Ð°Ð½Ð¸Ðµ Ð´Ð°Ð½Ð½Ñ‹Ñ…...")
    while True:
        try:
            bot.polling(none_stop=True, interval=0, timeout=30)
        except Exception as e:
            print(f"âš ï¸ ÐžÑˆÐ¸Ð±ÐºÐ°: {e}, Ð¿ÐµÑ€ÐµÐ·Ð°Ð¿ÑƒÑÐº Ñ‡ÐµÑ€ÐµÐ· 5 ÑÐµÐº...")
            time.sleep(5)
